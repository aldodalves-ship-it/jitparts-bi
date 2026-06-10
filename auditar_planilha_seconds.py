"""Audita a planilha ReportProfitability exportada da Seconds."""

from __future__ import annotations

import re
import unicodedata
from difflib import SequenceMatcher
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


INPUT_PATH = Path("data/seconds/ReportProfitability.xlsx")
TARGET_FATURAMENTO = 106_713.71

FINANCIAL_ALIASES = {
    "preco_venda": ["Preco da Venda", "Preço da Venda"],
    "total_faturado": ["Total_Faturado", "Total Faturado"],
    "vendidos": ["Vendidos"],
    "lucro_liquido": ["Lucro Liquido", "Lucro Líquido"],
    "lucro_bruto": ["Lucro Bruto"],
    "cmv": ["Custo da Mercadoria Vendida", "CMV"],
    "frete": ["Frete Gratis voce paga os custos", "Frete-Grátis (você paga os custos)"],
    "imposto": ["Imposto"],
}


def normalize_column_name(column: object) -> str:
    text = str(column).strip().lower().replace("\ufeff", "").replace("\ufffd", "")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def find_column(df: pd.DataFrame, aliases: list[str]) -> str | None:
    normalized_columns = {normalize_column_name(column): str(column) for column in df.columns}
    for alias in aliases:
        normalized_alias = normalize_column_name(alias)
        if normalized_alias in normalized_columns:
            return normalized_columns[normalized_alias]

    best_column: str | None = None
    best_score = 0.0
    for alias in aliases:
        normalized_alias = normalize_column_name(alias)
        for normalized_column, original_column in normalized_columns.items():
            score = SequenceMatcher(None, normalized_alias, normalized_column).ratio()
            if score > best_score:
                best_score = score
                best_column = original_column
    return best_column if best_score >= 0.86 else None


def parse_br_number(value: object) -> float:
    if pd.isna(value):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return 0.0

    text = (
        text.replace("R$", "")
        .replace("%", "")
        .replace("\u00a0", "")
        .replace(" ", "")
    )
    text = re.sub(r"[^0-9,.\-]", "", text)
    if not text or text in {"-", ".", ","}:
        return 0.0

    if "," in text and "." in text and text.rfind(",") > text.rfind("."):
        text = text.replace(".", "").replace(",", ".")
    elif "," in text and "." not in text:
        text = text.replace(",", ".")
    elif "," in text and "." in text and text.rfind(".") > text.rfind(","):
        text = text.replace(",", "")

    try:
        return float(text)
    except ValueError:
        return 0.0


def format_number(value: float) -> str:
    return f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def sheet_metadata(path: Path) -> None:
    wb = load_workbook(path, read_only=False, data_only=True)
    print("=" * 100)
    print("METADADOS DO ARQUIVO")
    print("=" * 100)
    print(f"Arquivo: {path}")
    print(f"Abas: {wb.sheetnames}")
    for ws in wb.worksheets:
        hidden_rows = [idx for idx, dim in ws.row_dimensions.items() if dim.hidden]
        hidden_columns = [key for key, dim in ws.column_dimensions.items() if dim.hidden]
        print(
            f"- Aba: {ws.title} | estado={ws.sheet_state} | linhas={ws.max_row - 1} "
            f"| colunas={ws.max_column} | auto_filter={ws.auto_filter.ref}"
        )
        print(f"  Linhas ocultas: {len(hidden_rows)}")
        print(f"  Colunas ocultas: {hidden_columns if hidden_columns else 'nenhuma'}")
        print(f"  Tabelas: {list(ws.tables.keys()) if ws.tables else 'nenhuma'}")


def numeric_columns(df: pd.DataFrame) -> dict[str, pd.Series]:
    columns: dict[str, pd.Series] = {}
    for label, aliases in FINANCIAL_ALIASES.items():
        column = find_column(df, aliases)
        if column is not None:
            columns[label] = df[column].map(parse_br_number)
    return columns


def print_sheet_audit(sheet_name: str, df: pd.DataFrame) -> None:
    print("\n" + "=" * 100)
    print(f"ABA: {sheet_name}")
    print("=" * 100)
    print(f"Linhas pandas: {len(df)}")
    print("Colunas:")
    print(list(df.columns))

    columns = numeric_columns(df)
    if not columns:
        print("Nenhuma coluna financeira conhecida encontrada.")
        return

    vendidos = columns.get("vendidos", pd.Series([0.0] * len(df), index=df.index))
    sold_mask = vendidos > 0

    print("\nSomas gerais:")
    for label, series in columns.items():
        print(f"- {label}: {format_number(float(series.sum()))}")

    print("\nProdutos com Vendidos > 0:")
    print(f"- Quantidade de produtos: {int(sold_mask.sum())}")
    print(f"- Soma Vendidos: {format_number(float(vendidos[sold_mask].sum()))}")
    for label, series in columns.items():
        print(f"- {label}: {format_number(float(series[sold_mask].sum()))}")

    print("\nColunas unitarias multiplicadas por Vendidos:")
    for label, series in columns.items():
        if label == "vendidos":
            continue
        multiplied = series * vendidos
        print(f"- {label} * Vendidos: {format_number(float(multiplied.sum()))}")

    print("\nPossiveis colunas/expressoes proximas ao faturamento visual da Seconds:")
    candidates: list[tuple[str, float, float]] = []
    for label, series in columns.items():
        if label == "vendidos":
            continue
        values = {
            f"{label} soma geral": float(series.sum()),
            f"{label} apenas Vendidos > 0": float(series[sold_mask].sum()),
            f"{label} * Vendidos": float((series * vendidos).sum()),
        }
        for name, value in values.items():
            candidates.append((name, value, abs(value - TARGET_FATURAMENTO)))

    for name, value, diff in sorted(candidates, key=lambda item: item[2])[:12]:
        print(f"- {name}: {format_number(value)} | dif={format_number(diff)}")

    item_column = find_column(df, ["Codigo do Anuncio", "Código do Anúncio", "MLB"])
    if item_column:
        item_ids = df[item_column].astype(str).str.strip()
        duplicated = item_ids[item_ids.ne("")].duplicated().sum()
        print("\nDuplicidade por Codigo do Anuncio:")
        print(f"- Duplicados: {int(duplicated)}")

    possible_total_rows = df[
        df.astype(str).apply(
            lambda row: row.str.contains("total|subtotal|geral", case=False, regex=True).any(),
            axis=1,
        )
    ]
    print("\nLinhas que parecem total/subtotal:")
    if possible_total_rows.empty:
        print("- Nenhuma encontrada.")
    else:
        print(possible_total_rows.head(20).to_string(index=False))


def main() -> None:
    if not INPUT_PATH.exists():
        raise FileNotFoundError(f"Arquivo nao encontrado: {INPUT_PATH}")

    sheet_metadata(INPUT_PATH)
    sheets = pd.read_excel(INPUT_PATH, sheet_name=None, engine="openpyxl")
    for sheet_name, df in sheets.items():
        print_sheet_audit(sheet_name, df)


if __name__ == "__main__":
    main()
